from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, colors, Alignment
import os
import sys

def get_data(engine):
    logdir = '/opt/New_emerging_SQL_comparative_analysis/sundp/'
    result = logdir + engine  + '/times.csv'
    time = []
    print result
    with open(result) as f:
        next(f)
        for line in f:
            line = line.replace('\n','').split("|")
            t = line[2].replace('000','')
            if (line[10] == 'FAILED'):
                time.append(t+"_failed")
            else:
                t = long(float(t))
                time.append(t)


    file_xlsx = '/opt/New_emerging_SQL_comparative_analysis/sundp/benchmark_auto1.xlsx'
    wb = load_workbook(file_xlsx)
    sheet = wb.get_sheet_by_name('hos')
    col = sheet.max_column+1
    cell = sheet.cell(row=1,column=col)
    col_index = cell.column
    # col_dim = sheet.column_dimensions[col_index]
    left, right, top, bottom = [Side(style='thin')]*4
    border = Border(left=left, right=right, top=top, bottom=bottom)
    wrap_alignment = Alignment(wrap_text=True,horizontal='center', vertical='center')
    cell.value = engine
    cell.border = border
    cell.alignment = wrap_alignment
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in range(0,len(time)):
        cell = sheet.cell(row=row+2,column=col)
        cell.value = time[row]
        cell.alignment = center_alignment
        cell.border = border
    end_row = str(row+2)
    sum = str(row+3)
    cell_sum = sheet[col_index+sum]
    cell_sum.value = '=SUM('+col_index+'2:'+col_index+end_row+')'
    cell_sum.alignment = center_alignment
    cell_sum.border = border
    geo = str(row+4)
    cell_geo = sheet[col_index+geo]
    cell_geo.value = '=GEOMEAN('+col_index+'2:'+col_index+end_row+')'
    cell_geo.alignment = center_alignment
    cell_geo.border = border

    wb.save(file_xlsx)

if __name__ == '__main__':
    args = sys.argv
    if len(args) > 1:
        engine = args[1]
        #date_dir = args[2]
        get_data(engine)
    else:
        print "lost argv"
